import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import subprocess

# 查询所有老师数据
query = """
SELECT 
  id, 
  name, 
  nickname, 
  phone, 
  email, 
  wechat,
  roles,
  teacherAttribute, 
  customerType, 
  category,
  hourlyRate,
  bankAccount,
  bankName,
  avatarUrl,
  isActive,
  createdAt,
  updatedAt
FROM users 
WHERE roles LIKE '%teacher%' AND deletedAt IS NULL
ORDER BY name ASC;
"""

# 执行SQL查询（通过MySQL命令行）
import os
db_url = os.environ.get('DATABASE_URL', '')
# 解析DATABASE_URL: mysql://user:pass@host:port/dbname
if db_url.startswith('mysql://'):
    db_url = db_url.replace('mysql://', '')
    if '@' in db_url:
        auth, rest = db_url.split('@', 1)
        user, password = auth.split(':', 1)
        if '/' in rest:
            host_port, dbname = rest.split('/', 1)
            if ':' in host_port:
                host, port = host_port.split(':', 1)
            else:
                host, port = host_port, '3306'
        else:
            host_port = rest
            dbname = 'course_crm'
            if ':' in host_port:
                host, port = host_port.split(':', 1)
            else:
                host, port = host_port, '3306'
    
    # 使用mysql命令行执行查询并输出JSON
    mysql_cmd = f'mysql -h {host} -P {port} -u {user} -p{password} {dbname} -e "{query}" --batch --skip-column-names'
    result = subprocess.run(mysql_cmd, shell=True, capture_output=True, text=True)
    
    # 解析结果
    teachers = []
    for line in result.stdout.strip().split('\n'):
        if line:
            fields = line.split('\t')
            if len(fields) >= 17:
                teachers.append({
                    'id': fields[0],
                    'name': fields[1] if fields[1] != 'NULL' else '',
                    'nickname': fields[2] if fields[2] != 'NULL' else '',
                    'phone': fields[3] if fields[3] != 'NULL' else '',
                    'email': fields[4] if fields[4] != 'NULL' else '',
                    'wechat': fields[5] if fields[5] != 'NULL' else '',
                    'roles': fields[6] if fields[6] != 'NULL' else '',
                    'teacherAttribute': fields[7] if fields[7] != 'NULL' else '',
                    'customerType': fields[8] if fields[8] != 'NULL' else '',
                    'category': fields[9] if fields[9] != 'NULL' else '',
                    'hourlyRate': fields[10] if fields[10] != 'NULL' else '',
                    'bankAccount': fields[11] if fields[11] != 'NULL' else '',
                    'bankName': fields[12] if fields[12] != 'NULL' else '',
                    'avatarUrl': fields[13] if fields[13] != 'NULL' else '',
                    'isActive': fields[14] if fields[14] != 'NULL' else '',
                    'createdAt': fields[15] if fields[15] != 'NULL' else '',
                    'updatedAt': fields[16] if fields[16] != 'NULL' else '',
                })

# 创建Excel工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "老师数据更新模板"

# 定义列标题
headers = [
    ("ID", "用户ID（不可修改）"),
    ("姓名", "老师姓名"),
    ("昵称", "昵称/别名"),
    ("手机号", "联系电话"),
    ("邮箱", "电子邮箱"),
    ("微信号", "微信号"),
    ("角色", "用户角色（不建议修改）"),
    ("老师属性", "老师类型/属性"),
    ("客户类型", "客户分类"),
    ("类别", "业务类别"),
    ("时薪", "课时费率"),
    ("银行账户", "收款账号"),
    ("开户行", "开户银行"),
    ("头像URL", "头像图片链接"),
    ("激活状态", "是否激活（1=是，0=否）"),
    ("创建时间", "账号创建时间（不可修改）"),
    ("更新时间", "最后更新时间（不可修改）"),
]

# 写入标题行
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col_idx, (header, _) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 写入说明行
desc_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
desc_font = Font(italic=True, size=9, color="404040")

for col_idx, (_, desc) in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx)
    cell.value = desc
    cell.fill = desc_fill
    cell.font = desc_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 写入数据行
for row_idx, teacher in enumerate(teachers, 3):
    ws.cell(row=row_idx, column=1, value=teacher['id'])
    ws.cell(row=row_idx, column=2, value=teacher['name'])
    ws.cell(row=row_idx, column=3, value=teacher['nickname'])
    ws.cell(row=row_idx, column=4, value=teacher['phone'])
    ws.cell(row=row_idx, column=5, value=teacher['email'])
    ws.cell(row=row_idx, column=6, value=teacher['wechat'])
    ws.cell(row=row_idx, column=7, value=teacher['roles'])
    ws.cell(row=row_idx, column=8, value=teacher['teacherAttribute'])
    ws.cell(row=row_idx, column=9, value=teacher['customerType'])
    ws.cell(row=row_idx, column=10, value=teacher['category'])
    ws.cell(row=row_idx, column=11, value=teacher['hourlyRate'])
    ws.cell(row=row_idx, column=12, value=teacher['bankAccount'])
    ws.cell(row=row_idx, column=13, value=teacher['bankName'])
    ws.cell(row=row_idx, column=14, value=teacher['avatarUrl'])
    ws.cell(row=row_idx, column=15, value=teacher['isActive'])
    ws.cell(row=row_idx, column=16, value=teacher['createdAt'])
    ws.cell(row=row_idx, column=17, value=teacher['updatedAt'])

# 调整列宽
column_widths = [10, 12, 12, 15, 25, 15, 20, 15, 12, 12, 10, 20, 20, 40, 10, 20, 20]
for col_idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# 冻结前两行
ws.freeze_panes = "A3"

# 保存Excel文件
output_file = "/home/ubuntu/老师数据更新模板.xlsx"
wb.save(output_file)
print(f"Excel模板已生成: {output_file}")
print(f"共包含 {len(teachers)} 位老师的数据")
